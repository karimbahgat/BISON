'''
Downloads the latest available shapefiles/zipfiles for each country
and automatically generates the sourceMetaData.json file.

NOTE: the dates on the website and filenames are the start dates of temporal validity, not the end dates.
Metadata for year and source_updated has to be determined based on the historical change table
on the country pages.
'''


import urllib.request
from zipfile import ZipFile
import os
import io
import re
import sys
import json
import logging
import datetime
import shapefile

# redirect to logfile
logger = sys.stdout #open('download_log.txt', mode='w', encoding='utf8', buffering=1)
#sys.stdout = logger
#sys.stderr = logger

# access the gadm country download page
root = 'https://www.unsalb.org'

def parse_country_links(raw):
    # hacky parse the html into elements
    elems = raw.replace('>','<').split('<')
    elems = (elem for elem in elems)
    elem = next(elems)

    # get all country links
    for elem in elems:
        # country link
        if elem.startswith('a href="/en/data/'):
            relUrl = elem.split(' ')[1].replace('href=', '').strip('"')
            url = root + relUrl
            #print('found country url', url)

        # check that has geospatial data
        if elem.startswith('td headers="view-datasets-count-table-column"'):
            #print('geospatial?')
            while elem != '/td':
                elem = next(elems)
                maybenum = elem.replace('\n','').strip(r'" ')
                if maybenum.isdigit() and maybenum != '0':
                    # count of geospatial data > 0
                    print('is geospatial!', url)
                    yield url

def iter_country_page_downloads(url):
    raw = urllib.request.urlopen(url).read().decode('utf8')

    # hacky parse the html into elements
    elems = raw.replace('>','<').split('<')
    elems = (elem for elem in elems)
    elem = next(elems)

    # find all shapefile downloads
    links = []
    for elem in elems:
        # find temporal info
        if 'temporal validity' in elem.lower():
            temporal_text = elem
            dates = [d for d in re.findall(r'\d\d\d\d-\d\d-\d\d', temporal_text)]
            if len(dates) != 2:
                print(f'validity text {temporal_text} must contain exactly 2 dates, not {dates}')
                continue

            dates = [d.replace('-00','-01') for d in dates] # some days/months are erroneously set to 00
            dates = [datetime.date.fromisoformat(d) for d in dates]
            dates = sorted(dates)
            fromdate,todate = dates
            print('FROM DATE',fromdate,'TO DATE',todate)

        # TODO: find source agency name and url
        # ... 

        # find shapefile download link
        if elem.startswith('a class="btn btn-primary mr-1"'):
            print('found download button')
            linktag = elem
            # find the text that indicates the type of file
            while elem != '/a':
                elem = next(elems)
                if 'shapefile' in elem.lower():
                    print('shapefile download')
                    # get the actual link
                    for part in linktag.split(' '):
                        if part.startswith('href='):
                            link = part.replace('href=','').strip('"')
                            print('FILE', link)
                            if link in links:
                                print('duplicate link, skipping')
                                continue
                            links.append(link)
                            yield fromdate,todate,link

def get_historical_table_download(url):
    raw = urllib.request.urlopen(url).read().decode('utf8')

    # hacky parse the html into elements
    elems = raw.replace('>','<').split('<')
    elems = (elem for elem in elems)
    elem = next(elems)

    # find all excel downloads
    downloads = []
    for elem in elems:
        elem = elem.replace('"','').replace("'","")
        if elem.endswith('.xlsx'):
            _url = elem.split('href=')[-1]
            if 'salb.org' not in _url:
                _url = f'{root}/{_url}' # from relative to abs url
            downloads.append(_url)
    if len(downloads) > 1:
        logging.warning('More than one historical tables, use only the first one')

    # return first download
    return downloads[0] if downloads else None

def parse_historical_table(url):
    from openpyxl import load_workbook
    wb = load_workbook(filename=io.BytesIO(urllib.request.urlopen(url).read()))
    # get metadata sheet
    metasheet = None
    for name in wb.sheetnames:
        if name.lower().startswith('metadata'):
            metasheet = wb[name]
            break
    # otherwise return early
    if not metasheet:
        logging.warning('Historal table had no metadata sheet')
        return {}
    # extract values from key-value row entries
    info = {}
    for row in metasheet.iter_rows():
        for i,cell in enumerate(row):
            if cell.value == 'Last update':
                updated = row[i+1].value
                info['updated'] = updated
    return info
        
def main():
    from adminManager import models
    from adminImporter.models import DatasetImporter

    from django.db import transaction
    from django.utils import timezone

    from . import utils

    with transaction.atomic():

        # get top source
        root_source = models.AdminSource.objects.get(name='UN SALB')

        # loop pages and download+unzip each
        print('begin')
        page = 0 # starts at page 0
        while True:
            url = '{}/data?page={}'.format(root, page)
            print('')
            print('looping country links from page', url)
            resp = urllib.request.urlopen(url)
            if resp.getcode() != 200:
                # reached the end/invalid page
                break

            # get country links
            raw = resp.read().decode('utf8')
            countrylinks = list(parse_country_links(raw))
            if len(countrylinks) == 0:
                # reached the end/invalid page
                break
            
            # loop
            for countrylink in countrylinks:
                print('countrylink', countrylink)
                
                # get page downloads
                page_downloads = list(iter_country_page_downloads(countrylink))
                if len(page_downloads) == 0:
                    print('No page downloads, skipping')
                    continue

                # create country folder
                # iso = countrylink[-3:].upper()
                # if os.path.lexists(iso):
                #     if REPLACE == False:
                #         # dont overwrite any existing folders
                #         # to protect any metadata that may have been manually edited
                #         print('Country folder already exists, skipping')
                #         continue
                # else:
                #     os.mkdir(iso)

                # create country source
                iso = countrylink[-3:].upper()
                meta = {
                    'parent': root_source,
                    'name': f'{iso}',
                    'url': countrylink,
                    'type': 'DataSource',
                }
                print(meta)
                cosource = models.AdminSource(**meta)
                cosource.save()

                # TODO: parse and create source lineage
                # ... 

                # loop file downloads
                for fromdate,todate,ziplink in page_downloads:
                    
                    # download zipfile
                    zipname = ziplink.split('/')[-1]
                    #if DOWNLOAD:
                    #    urllib.request.urlretrieve(ziplink, '{}/{}'.format(iso, zipname))
                    
                    # determine main entries for meta file
                    zip_contents = list(utils.inspect_zipfile_contents(ziplink))
                    shapefiles = [name
                                for name in zip_contents
                                if name.endswith('.shp')]
                    if len(shapefiles) == 0:
                        logging.warning('Zipfile does not contain any shapefiles: {}'.format(shapefiles))
                        continue

                    # in most cases this appears to be two files 'BNDA_*' and 'BNDL_*'
                    # it appears the one we want is 'BNDA_*'
                    bnd_shapefiles = [name for name in shapefiles
                                      if name.startswith(('BNDA_','SUBA_'))]
                    if len(bnd_shapefiles) == 0:
                        logging.warning('Zipfile does not contain any polygon shapefiles: {}'.format(shapefiles))
                        continue
                    
                    # lets assume there's only one shapefile left
                    shapefile_name = bnd_shapefiles[0]
                    shapefile_path = '{}/{}'.format(ziplink, shapefile_name)

                    # if present, use download link's "to" date text as basis for source_date
                    # in most cases the todate is set to "last update"
                    # but in cases where the shapefile has not been uploaded yet
                    # we can use the todate of the current shapefile
                    updated = None
                    if todate:
                        updated = todate
                        print('''source_update set to download link's "to" date text''')

                    # source_date fallback on historical table field if date/year hasn't been found yet
                    if not updated:
                        # try to find date info from historical excel table
                        table_download = get_historical_table_download(countrylink)
                        table_updated = None
                        if table_download:
                            table_info = parse_historical_table(table_download)
                            table_updated = table_info.get('updated', None)
                            if table_updated:
                                print('detected historical excel table "Last updated" field: {}'.format(table_updated))
                        else:
                            logging.warning('could not find historical table download')

                        if table_updated is None:
                            pass
                        elif hasattr(table_updated, 'year'):
                            # already date or datetime value
                            yr,mn,dy = table_updated.year, table_updated.month, table_updated.day
                            updated = datetime.date(yr, mn, dy)
                        else:
                            # date string
                            yr,mn,dy = table_updated.split('-')
                            dy,mn,yr = map(int, [dy,mn,yr])
                            updated = datetime.date(yr, mn, dy)
                        print('source_update set to historical excel table "Last updated" field')

                    # compare with "DATSOR" field in shapefile
                    # datsor = None
                    # if 'DATSOR' in fieldnames:
                    #     for rec in reader.iterRecords():
                    #         datsor = rec['DATSOR']
                    #         if datsor and len(datsor.split('/')) == 3:
                    #             break
                    #     if datsor:
                    #         print('detected "DATSOR" field in shapefile: {}'.format(datsor))
                        
                    # # fallback on datsor field if date/year hasn't been found yet
                    # if not updated and datsor:
                    #     dy,mn,yr = datsor.split('/')
                    #     dy,mn,yr = map(int, [dy,mn,yr])
                    #     updated = datetime.date(yr, mn, dy)
                    #     print('source_update set to shapefile "DATSOR" field')

                    # if nothing else, fallback on download link's "from" date text
                    # if not updated:
                    #     updated = fromdate
                    #     logging.warning('''source_update and year set to download link's "from" date text''')

                    # if missing todate, set to last updated
                    # if not todate:
                    #     todate = updated

                    # create timeperiod source
                    name = fromdate.isoformat() if fromdate else '?'
                    name += ' to '
                    name += todate.isoformat() if todate else '?'
                    meta = {
                        "parent":cosource,
                        "name":name,
                        "valid_from":fromdate.isoformat() if fromdate else None,
                        "valid_to":todate.isoformat() if todate else None,
                        #"source":["UN SALB"],
                        #"updated":updated.isoformat(),
                        "url":countrylink,
                        "type":"DataSource",
                    }
                    print(meta)
                    src = models.AdminSource(**meta)
                    src.save()

                    # create separate importers by dissolving each level
                    # most salb files are ADM2 indicated by ADM1NM and ADM2NM fields
                    # in rare cases, ADM2NM will be empty, indicating an adm1 boundary
                    max_level = 2
                    for level in range(max_level+1):
                        # create importer
                        import_params = {
                            'path':ziplink,
                            'path_ext':'.zip',
                            'path_zipped_file':shapefile_name,
                            'encoding':'utf8',
                            'levels':[
                                {'level':_lev,
                                'id_field':'adm{}cd'.format(_lev) if _lev > 0 else 'iso3cd',
                                'name_field':'adm{}nm'.format(_lev) if _lev > 0 else None,
                                #'codes':[{'type':'ISO 3166-1 alpha-3', 'value':iso if _lev==0 else None}]
                                }
                                for _lev in range(level+1)
                            ]
                        }
                        #import_params['levels'][0]['id'] = iso # set adm0 id to country iso
                        print(import_params)

                        importer = DatasetImporter(
                            source=src,
                            import_params=import_params,
                            import_status='Pending',
                            status_updated=timezone.now(),
                        )
                        importer.save()

            page += 1
            